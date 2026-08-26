"""
The daily .xlsx the dashboard offers for download.

A spreadsheet is not a CSV with a different extension. The point of handing
someone Excel rather than comma-separated text is that the file arrives ready
to work with: times that sort as times, counts that add up, a header that stays
put when you scroll, and the rows that need attention visible without reading
every one. That is what this builds.
"""

from __future__ import annotations

from datetime import date, datetime, time
from typing import Any, Dict, List, Optional, Sequence

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:                                   # pragma: no cover
    Workbook = None


HEADERS = (
    ("Clinic", 26),
    ("Date", 12),
    ("Opening time", 14),
    ("Closing time", 14),
    ("Status", 12),
    ("Offline minutes", 16),
    ("Checks", 9),
)

# Excel's own palette conventions: red for a problem, amber for a caveat,
# green for fine. Kept pale so the text stays readable when printed.
_FILL = {
    "opened": PatternFill("solid", fgColor="E8F5E9") if Workbook else None,
    "closed": PatternFill("solid", fgColor="FFF4E5") if Workbook else None,
    "offline": PatternFill("solid", fgColor="FDE7E9") if Workbook else None,
}
_TEXT = {
    "opened": "1B5E20",
    "closed": "8A5300",
    "offline": "B3261E",
}


def _as_time(value: str) -> Optional[time]:
    """"08:42" -> a real time, so Excel sorts and formats it as one."""
    if not value:
        return None
    try:
        hour, minute = (int(part) for part in value.split(":")[:2])
        return time(hour, minute)
    except (ValueError, TypeError):
        return None


def _as_date(value: str) -> Optional[date]:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def build_workbook(day: str, rows: Sequence[Dict[str, Any]]) -> "Workbook":
    """A one-sheet workbook for a single day, formatted and ready to read."""
    if Workbook is None:
        raise RuntimeError(
            "openpyxl is not installed on this host, so the Excel download is "
            "unavailable - install it with 'pip install openpyxl'"
        )

    book = Workbook()
    sheet = book.active
    sheet.title = day                      # e.g. "2026-08-25"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="37474F")
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centre = Alignment(horizontal="center", vertical="center")

    for index, (title, width) in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = border
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 26

    for offset, row in enumerate(rows):
        line = offset + 2
        status = row.get("status", "")
        values = (
            row.get("clinic", ""),
            _as_date(row.get("date", "")) or row.get("date", ""),
            _as_time(row.get("opening_time", "")),
            _as_time(row.get("closing_time", "")),
            status,
            int(row.get("offline_minutes") or 0),
            int(row.get("checks") or 0),
        )
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=line, column=index, value=value)
            cell.border = border
            if index == 2:
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = centre
            elif index in (3, 4):
                # A blank means nobody was ever seen, which is information; an
                # empty cell says it more clearly than "00:00" would.
                cell.number_format = "hh:mm"
                cell.alignment = centre
            elif index in (6, 7):
                cell.number_format = "0"
                cell.alignment = centre
            elif index == 5:
                cell.alignment = centre
                if _FILL.get(status):
                    cell.fill = _FILL[status]
                    cell.font = Font(bold=True, color=_TEXT.get(status, "000000"))

    last = len(rows) + 1
    if rows:
        # A real Excel table: banded rows, and every column filterable and
        # sortable from the header without anyone setting that up by hand.
        span = f"A1:{get_column_letter(len(HEADERS))}{last}"
        table = Table(displayName="DailyReport", ref=span)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1", showRowStripes=True, showColumnStripes=False
        )
        sheet.add_table(table)

    sheet.freeze_panes = "A2"              # header stays put while scrolling

    notes = book.create_sheet("What the columns mean")
    for line in _legend(day, rows):
        notes.append(line)
    notes.column_dimensions["A"].width = 20
    notes.column_dimensions["B"].width = 92
    for cell in notes["A"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")
    for cell in notes["B"]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    return book


def _legend(day: str, rows: Sequence[Dict[str, Any]]) -> List[Sequence[str]]:
    """
    The second sheet, so a figure is never read as more than it is.

    Whoever opens this weeks from now will not have been in the conversation
    where the caveats were explained, and a spreadsheet invites being trusted
    to the minute.
    """
    counts = {state: sum(1 for r in rows if r.get("status") == state)
              for state in ("opened", "closed", "offline")}
    return [
        ("Daily report", day),
        ("Clinics", f"{len(rows)} - {counts['opened']} opened, "
                    f"{counts['closed']} closed, {counts['offline']} offline"),
        ("", ""),
        ("Status", ""),
        ("opened", "Somebody was seen inside the clinic, so it was working."),
        ("closed", "The cameras worked and nobody was ever seen."),
        ("offline", "The device could not be reached, so nothing was watched. "
                    "This is NOT the same as closed - a clinic whose NVR drops "
                    "off the network looks identical to a shut one here, and "
                    "the fault is the connection, not the staff."),
        ("", ""),
        ("Opening time", "The first time a person was seen, not the moment the "
                         "door opened. The clinic may well have been working "
                         "earlier: somebody sitting still in a dim room is not "
                         "always picked up. Treat it as 'open by at least this "
                         "time'."),
        ("Closing time", "The last time a person was seen, read the same way."),
        ("Blank times", "Nobody was seen all day, or the device was offline."),
        ("Offline minutes", "How long the app reported the device unreachable. "
                            "Time inside this window was not watched at all, so "
                            "the opening and closing times cannot account for "
                            "it."),
        ("Checks", "How many patrol visits the row rests on. A day built from "
                   "3 checks is far weaker evidence than one built from 38 - "
                   "compare this before comparing times."),
        ("", ""),
        ("Cameras", "Times come from the indoor camera only. An outdoor view "
                    "cannot tell whether a clinic is working: an empty street "
                    "in the evening looks shut either way."),
    ]
